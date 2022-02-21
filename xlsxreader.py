import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook
thislist = []
thisdict = {}
directory = './excelsheets'
for filename in os.listdir(directory):
    f = os.path.join(directory, filename)
    if os.path.isfile(f):
        print(f)
        wb = load_workbook(filename = f)
        sheetname = wb.sheetnames
        sheet_ranges = wb[sheetname[0]]        
        aNum = 97
        column = chr(aNum)
        row = 5
        cellNum = f'{column}{row}'
        while sheet_ranges[cellNum].value is not None:
            thisdict["Invoice Number"] = sheet_ranges['f2'].value
            while sheet_ranges[cellNum].value is not None:
                header = f'{column}4'
                thisdict[f'{sheet_ranges[header].value}'] = sheet_ranges[cellNum].value
                aNum += 1
                column = chr(aNum)
                cellNum = f'{column}{row}'
            thislist.append(thisdict)
            thisdict = {}
            row += 1
            aNum = 97
            column = chr(aNum)
            cellNum = f'{column}{row}'
json_object = json.dumps(thislist, indent = 2, default = str)
print(json_object)
with open('person.json', 'w') as json_file:
    json.dump(thislist, json_file, indent = 2, default = str)

